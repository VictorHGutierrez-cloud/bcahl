import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_financial_statements_excel():
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Define styles
    header_font = Font(bold=True, size=12)
    section_font = Font(bold=True, size=11)
    data_font = Font(size=10)
    number_font = Font(size=10)
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    total_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    center_alignment = Alignment(horizontal="center", vertical="center")
    right_alignment = Alignment(horizontal="right", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # BALANCE SHEET - ASSETS
    ws_assets = wb.create_sheet("BALANCE_SHEET_ASSETS")
    
    # Headers
    ws_assets['A1'] = "AHL DESIGN & ENGINEERING SRL"
    ws_assets['A1'].font = Font(bold=True, size=16)
    ws_assets['A2'] = "BALANCE SHEET - ASSETS (ACTIF) 2023"
    ws_assets['A2'].font = Font(bold=True, size=14)
    
    # Period headers
    ws_assets['A4'] = "Period: 01/01/2023 to 31/12/2023"
    ws_assets['A4'].font = Font(bold=True)
    ws_assets['A5'] = "Period: 01/01/2022 to 31/12/2022"
    ws_assets['A5'].font = Font(bold=True)
    
    # Column headers
    headers = ['Description', 'Gross', 'Depr. & Prov.', '2023 Net', '2022 Net']
    for col, header in enumerate(headers, 1):
        cell = ws_assets.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Assets data based on original French text
    assets_data = [
        ['Subscribed capital not called (01)', '', '', '', ''],
        ['I. FIXED ASSETS (ACTIF IMMOBILISE)', '', '', '', ''],
        ['Establishment expenses', '', '', '', ''],
        ['Research and development', '', '', '', ''],
        ['Concessions, patents, trademarks, software and similar rights', 26334, 25736, 598, 13],
        ['Goodwill', '', '', '', ''],
        ['Other intangible assets', '', '', '', ''],
        ['Land', '', '', '', ''],
        ['Buildings', '', '', '', ''],
        ['Technical installations, equipment & industrial tools', 299209, 160970, 138239, 165982],
        ['Other tangible assets', '', '', '', ''],
        ['Assets under construction', '', '', '', ''],
        ['Advances & down payments', '', '', '', ''],
        ['Equity-accounted investments', '', '', '', ''],
        ['Other investments', '', '', '', ''],
        ['Receivables related to investments', '', '', '', ''],
        ['Other fixed securities', '', '', '', ''],
        ['Loans', '', '', '', ''],
        ['Other financial assets', 351854, 351854, 330711, ''],
        ['', '', '', '', ''],
        ['TOTAL (I)', 677397, 186706, 490691, 498020],
        ['', '', '', '', ''],
        ['II. CURRENT ASSETS (ACTIF CIRCULANT)', '', '', '', ''],
        ['Raw materials, supplies', '', '', '', ''],
        ['Goods in production', '', '', '', ''],
        ['Work in progress for services', '', '', '', ''],
        ['Intermediate and finished products', 705862, '', '', ''],
        ['Merchandise', '', '', '', ''],
        ['Advances & down payments paid on orders', 109370, 109370, 567648, ''],
        ['Customers and related accounts', 8779122, 623895, 8155227, 3586631],
        ['Other receivables', '', '', '', ''],
        ['Suppliers debtors', '', '', '', ''],
        ['Personnel', '', '', '', ''],
        ['Social organizations', 69444, 69444, '', ''],
        ['State, taxes on profits', 69203, 69203, 69203, ''],
        ['State, taxes on turnover', 1610060, 1610060, 2357072, ''],
        ['Others', 22414, 22414, 11096, ''],
        ['Subscribed and called-up capital, not paid', '', '', '', ''],
        ['Marketable securities', '', '', '', ''],
        ['Financial instruments and tokens held', '', '', '', ''],
        ['Cash and cash equivalents', 5462543, 5462543, 3216207, ''],
        ['Prepaid expenses', 226497, 226497, 77434, ''],
        ['', '', '', '', ''],
        ['TOTAL (II)', 16348653, 623895, 15724758, 10591153],
        ['', '', '', '', ''],
        ['Charges to be spread over several periods', '', '', '', ''],
        ['(III) Bond redemption premiums', '', '', '', ''],
        ['(IV) Conversion differences and asset evaluation differences', '', '', '', ''],
        ['(V)', '', '', '', '']
    ]
    
    # Fill assets data
    for row_idx, row_data in enumerate(assets_data, 8):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_assets.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if row_idx == 8:  # I. FIXED ASSETS header
                cell.font = section_font
                cell.fill = section_fill
            elif row_idx == 22:  # TOTAL (I)
                cell.font = section_font
                cell.fill = total_fill
            elif row_idx == 24:  # II. CURRENT ASSETS header
                cell.font = section_font
                cell.fill = section_fill
            elif row_idx == 44:  # TOTAL (II)
                cell.font = section_font
                cell.fill = total_fill
            else:
                cell.font = data_font
            
            if col_idx == 1:  # Description column
                cell.alignment = left_alignment
            else:  # Number columns
                cell.alignment = right_alignment
                if isinstance(value, (int, float)) and value != '':
                    cell.number_format = '#,##0'
    
    # Adjust column widths
    ws_assets.column_dimensions['A'].width = 50
    ws_assets.column_dimensions['B'].width = 15
    ws_assets.column_dimensions['C'].width = 15
    ws_assets.column_dimensions['D'].width = 15
    ws_assets.column_dimensions['E'].width = 15
    
    # LIABILITIES
    ws_liabilities = wb.create_sheet("BALANCE_SHEET_LIABILITIES")
    
    # Headers
    ws_liabilities['A1'] = "AHL DESIGN & ENGINEERING SRL"
    ws_liabilities['A1'].font = Font(bold=True, size=16)
    ws_liabilities['A2'] = "BALANCE SHEET - LIABILITIES (PASSIF) 2023"
    ws_liabilities['A2'].font = Font(bold=True, size=14)
    
    # Period headers
    ws_liabilities['A4'] = "Period: 01/01/2023 to 31/12/2023"
    ws_liabilities['A4'].font = Font(bold=True)
    ws_liabilities['A5'] = "Period: 01/01/2022 to 31/12/2022"
    ws_liabilities['A5'].font = Font(bold=True)
    
    # Column headers
    liab_headers = ['Description', '2023 Net', '2022 Net']
    for col, header in enumerate(liab_headers, 1):
        cell = ws_liabilities.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Liabilities data based on original French text
    liabilities_data = [
        ['I. SHAREHOLDERS\' EQUITY (CAPITAUX PROPRES)', '', ''],
        ['Issue premiums, merger premiums, contribution premiums', '', ''],
        ['Revaluation differences', '', ''],
        ['Legal reserve', '', ''],
        ['Statutory or contractual reserves', '', ''],
        ['Regulatory reserves', '', ''],
        ['Other reserves', '', ''],
        ['Brought forward balance', '', ''],
        ['Result of the period', 5400, 1080],
        ['Investment grants', 4014845, 1452607],
        ['Regulatory provisions', 5473932, 848377],
        ['Income from participatory securities issues', '', ''],
        ['Conditional advances', 2535039, 6200374],
        ['', '', ''],
        ['TOTAL (I)', 581529, 573733],
        ['', '', ''],
        ['II. PROVISIONS FOR RISKS AND CHARGES', '', ''],
        ['Provisions for risks', '', ''],
        ['Provisions for charges', 2465, 10741517],
        ['', '', ''],
        ['TOTAL (II)', 16215449, ''],
        ['', '', ''],
        ['III. DEBT (EMPRUNTS ET DETTES)', '', ''],
        ['Convertible bonds', '', ''],
        ['Other convertible bonds', '', ''],
        ['Borrowings from credit institutions', '', ''],
        ['Borrowings', '', ''],
        ['Overdrafts, bank facilities', '', ''],
        ['Various financial borrowings and debts', '', ''],
        ['Others', 5400, 1080],
        ['Partners', 3625525, 389321],
        ['Advances & down payments received on orders', 4021326, 239564],
        ['Supplier debts and related accounts', 98810, 2415458],
        ['Tax and social debts', '', ''],
        ['Personnel', '', ''],
        ['Social organizations', 4012516, 212686],
        ['State, taxes on profits', 46668, 42145],
        ['State, taxes on turnover', 7067847, 11089173],
        ['State, guaranteed obligations and other taxes payable', '', ''],
        ['Debts on fixed assets and related accounts', '', ''],
        ['Other debts', '', ''],
        ['Financial instruments', '', ''],
        ['Prepaid income', '', ''],
        ['', '', ''],
        ['TOTAL (III)', '', ''],
        ['TOTAL (IV)', '', ''],
        ['Conversion differences and liability evaluation differences (V)', '', ''],
        ['', '', ''],
        ['TOTAL ASSETS (I to V)', '', '']
    ]
    
    # Fill liabilities data
    for row_idx, row_data in enumerate(liabilities_data, 8):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_liabilities.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if row_idx == 8:  # I. SHAREHOLDERS' EQUITY header
                cell.font = section_font
                cell.fill = section_fill
            elif row_idx == 15:  # TOTAL (I)
                cell.font = section_font
                cell.fill = total_fill
            elif row_idx == 17:  # II. PROVISIONS header
                cell.font = section_font
                cell.fill = section_fill
            elif row_idx == 21:  # TOTAL (II)
                cell.font = section_font
                cell.fill = total_fill
            elif row_idx == 23:  # III. DEBT header
                cell.font = section_font
                cell.fill = section_fill
            else:
                cell.font = data_font
            
            if col_idx == 1:  # Description column
                cell.alignment = left_alignment
            else:  # Number columns
                cell.alignment = right_alignment
                if isinstance(value, (int, float)) and value != '':
                    cell.number_format = '#,##0'
    
    # Adjust column widths
    ws_liabilities.column_dimensions['A'].width = 50
    ws_liabilities.column_dimensions['B'].width = 20
    ws_liabilities.column_dimensions['C'].width = 20
    
    # PROFIT AND LOSS STATEMENT
    ws_pnl = wb.create_sheet("PROFIT_LOSS_STATEMENT")
    
    # Headers
    ws_pnl['A1'] = "AHL DESIGN & ENGINEERING SRL"
    ws_pnl['A1'].font = Font(bold=True, size=16)
    ws_pnl['A2'] = "PROFIT AND LOSS STATEMENT (COMPTE DE RESULTAT) 2023"
    ws_pnl['A2'].font = Font(bold=True, size=14)
    
    # Period headers
    ws_pnl['A4'] = "Period: 01/01/2023 to 31/12/2023"
    ws_pnl['A4'].font = Font(bold=True)
    ws_pnl['A5'] = "Period: 01/01/2022 to 31/12/2022"
    ws_pnl['A5'].font = Font(bold=True)
    
    # Column headers
    pnl_headers = ['Description', '2023 Total', '2022 Total']
    for col, header in enumerate(pnl_headers, 1):
        cell = ws_pnl.cell(row=7, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # P&L data based on original French text
    pnl_data = [
        ['I. OPERATING INCOME', '', ''],
        ['Sales of merchandise', 119722, 28723],
        ['Sales of production', 45626970, 17968685],
        ['Net Turnover', 45746692, 17997408],
        ['Stock production', -705862, 705862],
        ['Immobilized production', '', ''],
        ['Operating subsidies', '', ''],
        ['Reversals of depreciation and provisions', '', ''],
        ['Other income', 28699, 25221],
        ['', '', ''],
        ['TOTAL OPERATING INCOME (I)', 45068929, 18728491],
        ['', '', ''],
        ['II. OPERATING EXPENSES', '', ''],
        ['Purchases of merchandise (including customs duties)', 166, ''],
        ['Stock variation (merchandise)', '', ''],
        ['Purchases of raw materials and other supplies', 9170811, 3590970],
        ['Stock variation (raw materials and other supplies)', '', ''],
        ['Other purchases and external charges', 17633929, 10523976],
        ['Taxes and similar payments', 29272, 22122],
        ['Wages and salaries', 13156204, 3892973],
        ['Social charges', 2823344, 91512],
        ['Depreciation on fixed assets', 58359, 44296],
        ['Provisions on fixed assets', '', ''],
        ['Provisions on current assets', '', ''],
        ['Provisions for risks and charges', '', ''],
        ['Other charges', 36396, 19553],
        ['', '', ''],
        ['TOTAL OPERATING EXPENSES (II)', 42908249, 18185568],
        ['', '', ''],
        ['OPERATING RESULT (I-II)', 2160680, 542923],
        ['', '', ''],
        ['III. FINANCIAL INCOME', '', ''],
        ['Equity share of results (joint ventures)', '', ''],
        ['Profit or loss transferred', 0, ''],
        ['Loss to be borne or transferred', 0, ''],
        ['Income from investment securities', '', ''],
        ['Income from other marketable securities and receivables', '', ''],
        ['Other interest and similar income', 14187, ''],
        ['Reversals of provisions and charge transfers', '', ''],
        ['Positive exchange differences', 420951, 261153],
        ['Net income on sales of marketable securities', '', ''],
        ['', '', ''],
        ['TOTAL FINANCIAL INCOME (IV)', 420985, 261340],
        ['', '', ''],
        ['IV. FINANCIAL EXPENSES', '', ''],
        ['Financial provisions for depreciation and provisions', '', ''],
        ['Interest and similar charges', 8289, 10733],
        ['Negative exchange differences', 506971, 333690],
        ['Charges on sales of marketable securities', '', ''],
        ['', '', ''],
        ['TOTAL FINANCIAL EXPENSES (VI)', 515260, 344423],
        ['', '', ''],
        ['FINANCIAL RESULT (V-VI)', -94275, -83083],
        ['', '', ''],
        ['CURRENT RESULT BEFORE TAXES', 1066405, 459840],
        ['', '', ''],
        ['V. EXCEPTIONAL INCOME', '', ''],
        ['Exceptional income on management operations', '', ''],
        ['Exceptional income on capital operations', '', ''],
        ['Reversals of provisions and charge transfers', '', ''],
        ['', '', ''],
        ['TOTAL EXCEPTIONAL INCOME (VII)', 0, 0],
        ['', '', ''],
        ['VI. EXCEPTIONAL EXPENSES', '', ''],
        ['Exceptional expenses on management operations', 187223, 918],
        ['Exceptional expenses on capital operations', '', ''],
        ['Exceptional provisions for depreciation and provisions', '', ''],
        ['', '', ''],
        ['TOTAL EXCEPTIONAL EXPENSES (VIII)', 187223, 918],
        ['', '', ''],
        ['EXCEPTIONAL RESULT (VII - VIII)', -187223, -918],
        ['', '', ''],
        ['Employee profit-sharing (IX)', '', ''],
        ['Taxes on profits (X)', 426577, 69601],
        ['', '', ''],
        ['TOTAL INCOME (I + III + VII)', 45489914, 18989831],
        ['TOTAL EXPENSES (II + IV + VI + VIII + IX + X)', 44037309, 18600510],
        ['', '', ''],
        ['NET RESULT', 1452605, 389321],
        ['Profit', 'Profit', ''],
        ['Including equipment leasing', '', ''],
        ['Including real estate leasing', '', '']
    ]
    
    # Fill P&L data
    for row_idx, row_data in enumerate(pnl_data, 8):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_pnl.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if row_idx in [8, 13, 34, 45, 55, 62]:  # Section headers
                cell.font = section_font
                cell.fill = section_fill
            elif row_idx in [11, 29, 32, 43, 47, 50, 60, 65, 68]:  # Totals
                cell.font = section_font
                cell.fill = total_fill
            else:
                cell.font = data_font
            
            if col_idx == 1:  # Description column
                cell.alignment = left_alignment
            else:  # Number columns
                cell.alignment = right_alignment
                if isinstance(value, (int, float)) and value != '':
                    cell.number_format = '#,##0'
    
    # Adjust column widths
    ws_pnl.column_dimensions['A'].width = 50
    ws_pnl.column_dimensions['B'].width = 20
    ws_pnl.column_dimensions['C'].width = 20
    
    # Save the workbook
    filename = "AHL_DESIGN_ENGINEERING_Financial_Statements_2023_COMPLETE.xlsx"
    wb.save(filename)
    print(f"Excel file created successfully: {filename}")
    return filename

if __name__ == "__main__":
    create_financial_statements_excel()
